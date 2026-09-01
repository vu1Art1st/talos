"""测试计划 Excel 导入导出：导出工作簿构建、导入模板构建与 upsert 导入。

列结构与导入/导出共用同一套表头；导入按 ID 更新、无 ID 新增，
工单ID重复时整批终止（与后端唯一性校验口径一致）。
"""
from fastapi import HTTPException
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.constants import TESTING_PLAN_STATUS, PlanStatus
from app.core.filters import to_float
from app.core.sanitize import excel_safe
from app.models import TestingPlan, User
from app.schemas import PlanImportResultOut
from app.services import ticket_service

PLAN_EXCEL_HEADERS = [
    "ID", "渗透测试工单名称", "测试系统", "测试类型", "所属部门",
    "工单ID", "工单提起时间", "状态", "测试人员",
    "需求接收", "初测完成", "复测通知", "复测完成",
    "预估人天", "实际人天",
    "超危数", "高危数", "中危数", "低危数", "复测轮数",
]

PLAN_STATUS_REVERSE = {v: k for k, v in TESTING_PLAN_STATUS.items()}


def _to_int(text: str) -> int:
    try:
        return int(float(text)) if text else 0
    except ValueError:
        return 0


def build_export_workbook(plans: list[TestingPlan], stats: dict) -> Workbook:
    """导出工作簿：明细 sheet（筛选后的工单行）+ 统计汇总 sheet。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "渗透测试工单"
    ws.append(PLAN_EXCEL_HEADERS)
    for p in plans:
        ws.append([excel_safe(v) for v in (
            p.id, p.plan_name, p.system_name, p.test_type, p.department,
            p.ticket_id, p.ticket_time,
            TESTING_PLAN_STATUS.get(p.status, str(p.status)),
            "、".join(u.realname or u.username for u in p.testers),
            p.receive_time, p.first_test_done_time, p.retest_notice_time, p.retest_done_time,
            p.est_mandays, p.actual_mandays,
            p.stat_critical, p.stat_high, p.stat_medium, p.stat_low,
            p.retest_round_count,
        )])

    ws2 = wb.create_sheet("统计汇总")
    ws2.append(["指标", "数值"])
    ws2.append(["渗透测试工单总数", stats["total_plans"]])
    ws2.append(["复测完成计划数", stats["retest_done_plans"]])
    ws2.append(["初测次数", stats["first_test_count"]])
    ws2.append(["复测次数", stats["retest_count"]])
    ws2.append(["总测试次数（初测+复测）", stats["total_test_count"]])
    ws2.append(["预估人天总计", stats["est_mandays_total"]])
    ws2.append(["实际人天总计", stats["actual_mandays_total"]])
    ws2.append(["剩余预估人天（未测试）", stats["remaining_est_mandays"]])
    ws2.append([])
    ws2.append(["状态", "计划数"])
    for row in stats["by_status"]:
        ws2.append([row["name"], row["count"]])
    ws2.append([])
    ws2.append(["月份", "漏洞数"])
    for row in stats["vulns_by_month"]:
        ws2.append([row["month"], row["count"]])
    return wb


def build_template_workbook() -> Workbook:
    """导入模板：列与导出一致，附一行示例数据（ID 留空则新增）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "渗透测试工单"
    ws.append(PLAN_EXCEL_HEADERS)
    ws.append([
        "", "示例渗透测试工单", "示例商城系统", "渗透测试", "电商事业部",
        "", "2026-01-01", "未测试", "张三、李四",
        "2026-01-01", "", "", "",
        5, 0,
        0, 0, 0, 0, 0,
    ])
    return wb


async def _load_user_map(session: AsyncSession) -> dict[str, User]:
    """测试人员匹配表（姓名优先，其次用户名），供导入行内姓名解析。"""
    users = (await session.execute(select(User))).scalars().all()
    user_map: dict[str, User] = {}
    for u in users:
        if u.realname:
            user_map.setdefault(u.realname, u)
        user_map.setdefault(u.username, u)
    return user_map


async def upsert_plans(session: AsyncSession, wb, user: User) -> PlanImportResultOut:
    """逐行导入测试计划：按 ID 更新、无 ID 新增，测试人员按姓名/用户名匹配。

    工单ID占用表（手动值或自动生成值均计入）用于批内与库内唯一性校验，
    重复时整批终止并提示，由调用方统一回滚。
    """
    user_map = await _load_user_map(session)
    occupied: dict[str, int | str] = {}
    for p in (await session.execute(select(TestingPlan))).scalars().all():
        tid = p.ticket_id
        if tid:
            occupied[tid] = p.id

    ws = wb.active
    result = PlanImportResultOut()
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = [str(c).strip() if c is not None else "" for c in row]
        cells += [""] * (len(PLAN_EXCEL_HEADERS) - len(cells))
        if not any(cells):
            continue
        result.total += 1
        system_name = cells[2]
        if not system_name:
            result.failed += 1
            result.errors.append(f"第{idx}行：测试系统为必填项")
            continue

        row_id = _to_int(cells[0])
        plan = await session.get(TestingPlan, row_id) if row_id else None
        is_new = plan is None
        if is_new:
            plan = TestingPlan(creator_id=user.id)
            session.add(plan)

        plan.plan_name = cells[1]
        plan.system_name = system_name
        plan.test_type = cells[3]
        plan.department = cells[4]
        # cells[5] 工单ID：显式填写则作为手动指定值，未填写则保持原值（新记录由系统自动生成）
        plan.ticket_id_manual = cells[5] or plan.ticket_id_manual or ""
        plan.ticket_time = cells[6]
        plan.status = PLAN_STATUS_REVERSE.get(cells[7], PlanStatus.UNTESTED)
        plan.receive_time = cells[9]
        plan.first_test_done_time = cells[10]
        plan.retest_notice_time = cells[11]
        plan.retest_done_time = cells[12]
        plan.est_mandays = to_float(cells[13])
        plan.actual_mandays = to_float(cells[14])
        plan.stat_critical = _to_int(cells[15])
        plan.stat_high = _to_int(cells[16])
        plan.stat_medium = _to_int(cells[17])
        plan.stat_low = _to_int(cells[18])
        matched = [user_map[name] for name in cells[8].split("、") if name.strip() and name.strip() in user_map]
        if matched:
            plan.testers = matched
        # 新增或历史数据无序号时按需求接收日期自动补号
        await ticket_service.assign_ticket_seq(session, plan)

        # 工单ID唯一性校验：与库中或批内其他行重复则整批终止并提示
        tid = plan.ticket_id
        if tid:
            if tid in occupied and occupied[tid] != row_id:
                raise HTTPException(400, f"第{idx}行：工单ID「{tid}」已存在，请更换后重新导入")
            occupied[tid] = "new" if is_new else row_id

        if is_new:
            result.created += 1
        else:
            result.updated += 1
    await session.commit()
    return result


# 结论性输出附件表头：工单ID / 所属部门 / 测试系统 / 漏洞数 / 测试类型 / 整改完成情况
CONCLUSION_HEADERS = ["工单ID", "所属部门", "测试系统", "漏洞数", "测试类型", "整改完成情况"]


def build_conclusion_workbook(rows: list[dict]) -> Workbook:
    """结论性输出附件：单 sheet，每行一个渗透测试工单的整改情况。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "整改情况"
    ws.append(CONCLUSION_HEADERS)
    for r in rows:
        ws.append([excel_safe(v) for v in (
            r.get("ticket_id", ""),
            r.get("department", ""),
            r.get("system_name", ""),
            r.get("vuln_count", 0),
            r.get("test_type", ""),
            r.get("rectify_state", ""),
        )])
    return wb
