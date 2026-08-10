"""专项管理 API：远程检测 / 测试计划 / 春耕行动，统一 special:manage 权限。"""
import html as html_mod
import json
import re
from datetime import datetime, timedelta
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import String, and_, exists, func, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.constants import NONPEN_ITEMS, TESTING_PLAN_STATUS, PlanStatus
from app.core.deps import require_any_perm, require_perm
from app.core.query import get_or_404, paginate, apply_sort
from app.core.sanitize import excel_safe
from app.core.timeutil import now as tznow
from app.db import get_session
from app.models import (
    Message,
    NonpenPlan,
    RemoteTesting,
    Report,
    ReportSection,
    SpringAction,
    TestingPlan,
    TestingPlanRetestRound,
    User,
    Vul,
    testing_plan_testers,
)
from app.schemas import (
    Page,
    PlanImportResultOut,
    RemoteTestingIn,
    RemoteTestingOut,
    SpringActionIn,
    SpringActionOut,
    TestingPlanIn,
    TestingPlanOut,
)
from app.services import nonpen_service, plan_service, ticket_service, vuln_service

router = APIRouter(tags=["专项管理"])


# ---------- 远程检测 ----------
@router.get("/remote-testings", response_model=Page[RemoteTestingOut])
async def list_remote_testings(
    search: str = "",
    sort: str = "",
    order: str = "desc",
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    _: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    cond = []
    if search:
        cond.append(
            RemoteTesting.title.ilike(f"%{search}%")
            | RemoteTesting.system_name.ilike(f"%{search}%")
            | RemoteTesting.department.ilike(f"%{search}%")
        )
    stmt = select(RemoteTesting).where(*cond)
    stmt = apply_sort(
        stmt, RemoteTesting, sort, order,
        {"id", "title", "system_name", "test_time", "department", "appeal_success", "create_time"},
        RemoteTesting.id.desc(),
    )
    total, items = await paginate(session, stmt, page, size)
    return Page(total=total, items=items)


async def _check_appeal_report(session: AsyncSession, report_id: int | None) -> None:
    if report_id is not None and await session.get(Report, report_id) is None:
        raise HTTPException(400, "指定的申诉报告不存在")


@router.post("/remote-testings", response_model=RemoteTestingOut)
async def create_remote_testing(
    body: RemoteTestingIn,
    user: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    await _check_appeal_report(session, body.appeal_report_id)
    row = RemoteTesting(**body.model_dump(), creator_id=user.id)
    session.add(row)
    await session.commit()
    await session.refresh(row)
    return row


@router.put("/remote-testings/{row_id}", response_model=RemoteTestingOut)
async def update_remote_testing(
    row_id: int,
    body: RemoteTestingIn,
    _: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    row = await get_or_404(session, RemoteTesting, row_id, "远程检测记录不存在")
    await _check_appeal_report(session, body.appeal_report_id)
    for k, v in body.model_dump().items():
        setattr(row, k, v)
    await session.commit()
    await session.refresh(row)
    return row


@router.delete("/remote-testings/{row_id}")
async def delete_remote_testing(
    row_id: int,
    _: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    row = await session.get(RemoteTesting, row_id)
    if row:
        await session.delete(row)
        await session.commit()
    return {"msg": "删除成功"}


# ---------- 测试计划 ----------
def _plan_conditions(
    search: str = "",
    status: int | None = None,
    test_type: str = "",
    department: str = "",
    receive_from: str = "",
    receive_to: str = "",
    tester_id: int | None = None,
    unclaimed: bool = False,
    pending: bool = False,
) -> list:
    """测试计划筛选条件，供列表/统计/导出共用。receive_time 为 YYYY-MM-DD 字符串，直接比较。

    tester_id 非空时过滤「当前可测试系统」：当前用户为测试人且状态为初测中/复测申请/复测中。
    unclaimed 为真时过滤「无人认领的测试」：测试人员列表为空。
    pending 为真时过滤「待办流程」：状态为未测试/初测中/复测中。
    两个快捷模式同时启用时按并集处理：满足任一条件的记录均展示。
    """
    cond = []
    if search:
        cond.append(
            TestingPlan.system_name.ilike(f"%{search}%")
            | TestingPlan.department.ilike(f"%{search}%")
            | TestingPlan.test_type.ilike(f"%{search}%")
        )
    if status is not None:
        cond.append(TestingPlan.status == status)
    if pending:
        # 待办流程：未测试 / 初测中 / 复测中
        cond.append(
            TestingPlan.status.in_([
                PlanStatus.UNTESTED, PlanStatus.TESTING, PlanStatus.RETESTING,
            ])
        )
    if test_type:
        cond.append(TestingPlan.test_type == test_type)
    if department:
        cond.append(TestingPlan.department == department)
    if receive_from:
        cond.append(TestingPlan.receive_time >= receive_from)
    if receive_to:
        # 空 receive_time 恒小于任意日期串，仅需上界比较时排除空值
        cond.append(TestingPlan.receive_time != "")
        cond.append(TestingPlan.receive_time <= receive_to)
    if tester_id is not None and unclaimed:
        # 并集：当前可测试系统 OR 无人认领（与其他筛选条件保持 AND）
        cond.append(
            or_(
                and_(
                    TestingPlan.status.in_([
                        PlanStatus.TESTING, PlanStatus.RETEST_APPLY, PlanStatus.RETESTING,
                    ]),
                    exists().where(
                        testing_plan_testers.c.testing_plan_id == TestingPlan.id,
                        testing_plan_testers.c.user_id == tester_id,
                    ),
                ),
                ~exists().where(testing_plan_testers.c.testing_plan_id == TestingPlan.id),
            )
        )
    elif tester_id is not None:
        cond.append(
            TestingPlan.status.in_([
                PlanStatus.TESTING, PlanStatus.RETEST_APPLY, PlanStatus.RETESTING,
            ])
        )
        cond.append(
            exists().where(
                testing_plan_testers.c.testing_plan_id == TestingPlan.id,
                testing_plan_testers.c.user_id == tester_id,
            )
        )
    elif unclaimed:
        cond.append(
            ~exists().where(testing_plan_testers.c.testing_plan_id == TestingPlan.id)
        )
    return cond


# ---------- 聚合筛选（filters JSON 参数） ----------
# 可筛选字段白名单：(列名, 字段类型, 是否为 DateTime 列)。日期字符串字段默认 "" 表示空，
# DateTime 列默认 NULL，数字字段默认 0。ticket_id / testers / *_count 为派生/关联字段，单独处理。
_PLAN_FILTER_FIELDS: dict[str, tuple[str, str, bool]] = {
    "id": ("id", "number", False),
    "plan_name": ("plan_name", "text", False),
    "system_name": ("system_name", "text", False),
    "test_type": ("test_type", "text", False),
    "department": ("department", "text", False),
    "status": ("status", "enum", False),
    "ticket_id": ("ticket_id", "text", False),  # 派生字段：手动指定或 receive_time+ticket_seq 生成
    "ticket_time": ("ticket_time", "date", False),
    "receive_time": ("receive_time", "date", False),
    "first_test_done_time": ("first_test_done_time", "date", False),
    "retest_notice_time": ("retest_notice_time", "date", False),
    "retest_done_time": ("retest_done_time", "date", False),
    "est_mandays": ("est_mandays", "number", False),
    "actual_mandays": ("actual_mandays", "number", False),
    "stat_critical": ("stat_critical", "number", False),
    "stat_high": ("stat_high", "number", False),
    "stat_medium": ("stat_medium", "number", False),
    "stat_low": ("stat_low", "number", False),
    "testers": ("testers", "text", False),  # 测试人员：多对多姓名/用户名匹配
    "vul_count": ("vul_count", "number", False),  # 关联漏洞计数
    "report_count": ("report_count", "number", False),  # 关联报告计数
    "retest_round_count": ("retest_round_count", "number", False),  # 复测轮数
    "create_time": ("create_time", "date", True),
    "update_time": ("update_time", "date", True),
}

_ALLOWED_FILTER_OPS = {
    "eq", "ne", "contains", "not_contains", "starts_with", "ends_with",
    "gt", "gte", "lt", "lte", "between", "is_empty", "is_not_empty",
}


def _split_range(value) -> tuple:
    """between 操作符取值：优先 [lo, hi] 数组，兼容 'lo,hi' 字符串。"""
    if isinstance(value, (list, tuple)) and len(value) == 2:
        return value[0], value[1]
    if isinstance(value, str) and "," in value:
        a, b = value.split(",", 1)
        return a, b
    raise HTTPException(400, "between 操作符需要两个值（如 [起始值, 结束值]）")


def _plan_filter_expr(col, ftype: str, is_datetime: bool, op: str, value) -> object:
    """按字段类型构造单字段筛选条件（不含 NOT 取反）。"""
    if op in ("is_empty", "is_not_empty"):
        if ftype == "number":
            empty = col.is_(None) | (col == 0)
        elif is_datetime:
            empty = col.is_(None)
        else:
            empty = col.is_(None) | (col == "")
        return empty if op == "is_empty" else ~empty
    if is_datetime:
        # DateTime 列统一转日期字符串比较，保证跨数据库行为一致
        col = func.date(col)
    if ftype in ("text", "enum"):
        if op == "contains":
            return col.ilike(f"%{value}%")
        if op == "not_contains":
            return ~col.ilike(f"%{value}%")
        if op == "starts_with":
            return col.ilike(f"{value}%")
        if op == "ends_with":
            return col.ilike(f"%{value}")
        if op == "eq":
            return col == value
        if op == "ne":
            return col != value
        raise HTTPException(400, f"文本字段不支持操作符：{op}")
    # number / date
    if op == "eq":
        return col == value
    if op == "ne":
        return col != value
    if op == "between":
        lo, hi = _split_range(value)
        if lo in (None, "") or hi in (None, ""):
            raise HTTPException(400, "区间筛选需要填写完整的起止值")
        if ftype == "number":
            return and_(col >= _to_float(lo), col <= _to_float(hi))
        # 日期字符串比较：排除空值
        return and_(col.is_not(None), col != "", col >= lo, col <= hi)
    if ftype == "number":
        v = _to_float(value) if not isinstance(value, (int, float)) else float(value)
        if op == "gt":
            return col > v
        if op == "gte":
            return col >= v
        if op == "lt":
            return col < v
        if op == "lte":
            return col <= v
        raise HTTPException(400, f"数字字段不支持操作符：{op}")
    # 日期字符串（YYYY-MM-DD 字典序即时间序）：上界比较需排除空值，下界比较天然排除空串
    if op in ("lt", "lte"):
        cond = col < value if op == "lt" else col <= value
        return and_(col.is_not(None), col != "", cond)
    if op == "gt":
        return col > value
    if op == "gte":
        return col >= value
    raise HTTPException(400, f"字段不支持操作符：{op}")


def _ticket_id_filter_expr(op: str, value) -> object:
    """工单ID为派生字段：手动指定值优先，否则由 receive_time(YYYY-MM-DD) + ticket_seq 生成 YYYYMMDD-N。"""
    sv = str(value) if value is not None else ""
    m = re.fullmatch(r"(\d{8})-(\d+)", sv)
    date_like = f"{m.group(1)[:4]}-{m.group(1)[4:6]}-{m.group(1)[6:]}%" if m else None

    def _eq():
        conds = [TestingPlan.ticket_id_manual == sv]
        if date_like:
            conds.append(and_(
                TestingPlan.receive_time.like(date_like),
                TestingPlan.ticket_seq == int(m.group(2)),
            ))
        return or_(*conds)

    if op == "eq":
        return _eq()
    if op == "ne":
        return ~_eq()
    if op in ("contains", "not_contains"):
        expr = or_(
            TestingPlan.ticket_id_manual.ilike(f"%{sv}%"),
            TestingPlan.receive_time.ilike(f"%{sv}%"),
            func.replace(TestingPlan.receive_time, "-", "").ilike(f"%{sv}%"),
            func.cast(TestingPlan.ticket_seq, String).ilike(f"%{sv}%"),
        )
        return expr if op == "contains" else ~expr
    if op == "starts_with":
        if len(sv) >= 8 and sv[:8].isdigit():
            d = f"{sv[:4]}-{sv[4:6]}-{sv[6:8]}"
            return or_(
                TestingPlan.ticket_id_manual.ilike(f"{sv}%"),
                TestingPlan.receive_time.like(f"{d}%"),
            )
        return or_(
            TestingPlan.ticket_id_manual.ilike(f"{sv}%"),
            TestingPlan.receive_time.like(f"{sv}%"),
        )
    if op == "ends_with":
        return or_(
            TestingPlan.ticket_id_manual.ilike(f"%{sv}"),
            func.cast(TestingPlan.ticket_seq, String).ilike(f"%{sv}"),
        )
    if op in ("is_empty", "is_not_empty"):
        empty = (TestingPlan.ticket_id_manual == "") & (TestingPlan.receive_time == "")
        return empty if op == "is_empty" else ~empty
    raise HTTPException(400, f"工单ID字段不支持操作符：{op}")


def _testers_filter_expr(op: str, value) -> object:
    """测试人员筛选：多对多关联 users 表，按姓名/用户名模糊匹配。"""
    if op in ("is_empty", "is_not_empty"):
        sub = exists().where(testing_plan_testers.c.testing_plan_id == TestingPlan.id)
        return ~sub if op == "is_empty" else sub
    sv = str(value) if value is not None else ""
    pat = f"%{sv}%"
    if op == "starts_with":
        pat = f"{sv}%"
    elif op == "ends_with":
        pat = f"%{sv}"
    sub = exists().where(
        testing_plan_testers.c.testing_plan_id == TestingPlan.id,
        testing_plan_testers.c.user_id == User.id,
        or_(User.realname.ilike(pat), User.username.ilike(pat)),
    )
    if op in ("eq", "contains", "starts_with", "ends_with"):
        return sub
    if op in ("ne", "not_contains"):
        return ~sub
    raise HTTPException(400, f"测试人员字段不支持操作符：{op}")


def _plan_count_expr(field: str, op: str, value) -> object:
    """关联计数筛选：关联漏洞数 / 关联报告数 / 复测轮数，通过相关子查询比较。"""
    if field == "vul_count":
        sub = select(func.count(Vul.id)).where(Vul.testing_plan_id == TestingPlan.id).scalar_subquery()
    elif field == "report_count":
        sub = select(func.count(Report.id)).where(Report.testing_plan_id == TestingPlan.id).scalar_subquery()
    else:
        sub = select(func.count(TestingPlanRetestRound.id)).where(
            TestingPlanRetestRound.plan_id == TestingPlan.id
        ).scalar_subquery()
    if op in ("is_empty", "is_not_empty"):
        empty = sub == 0
        return empty if op == "is_empty" else ~empty
    v = _to_float(value) if not isinstance(value, (int, float)) else float(value)
    if op == "eq":
        return sub == v
    if op == "ne":
        return sub != v
    if op == "gt":
        return sub > v
    if op == "gte":
        return sub >= v
    if op == "lt":
        return sub < v
    if op == "lte":
        return sub <= v
    if op == "between":
        lo, hi = _split_range(value)
        if lo in (None, "") or hi in (None, ""):
            raise HTTPException(400, "区间筛选需要填写完整的起止值")
        return and_(sub >= _to_float(lo), sub <= _to_float(hi))
    raise HTTPException(400, f"计数字段不支持操作符：{op}")


def _plan_filters_condition(filters: str) -> list:
    """解析聚合筛选 JSON（rules + 规则间 and/or 连接 + 单规则 not 取反），返回与现有条件 AND 的条件列表。

    请求格式示例：
        {"rules": [
            {"field": "status", "op": "eq", "value": 20, "not": false, "connector": "and"},
            {"field": "system_name", "op": "contains", "value": "商城", "not": true, "connector": "or"},
        ]}
    rules 按顺序组合：首条规则的 connector 忽略，其余规则的 connector 表示其与上一条之间的逻辑。
    """
    if not filters:
        return []
    try:
        payload = json.loads(filters)
    except (ValueError, TypeError):
        raise HTTPException(400, "filters 参数格式错误，需为 JSON 字符串")
    rules = payload.get("rules") if isinstance(payload, dict) else None
    if not isinstance(rules, list) or not rules:
        return []
    expr = None
    for r in rules:
        if not isinstance(r, dict):
            continue
        field = str(r.get("field", ""))
        op = str(r.get("op", ""))
        if op not in _ALLOWED_FILTER_OPS:
            raise HTTPException(400, f"不支持的操作符：{op}")
        if field not in _PLAN_FILTER_FIELDS:
            raise HTTPException(400, f"不支持的筛选字段：{field}")
        _, ftype, is_datetime = _PLAN_FILTER_FIELDS[field]
        if field == "ticket_id":
            cond = _ticket_id_filter_expr(op, r.get("value"))
        elif field == "testers":
            cond = _testers_filter_expr(op, r.get("value"))
        elif field in ("vul_count", "report_count", "retest_round_count"):
            cond = _plan_count_expr(field, op, r.get("value"))
        else:
            cond = _plan_filter_expr(
                getattr(TestingPlan, _PLAN_FILTER_FIELDS[field][0]), ftype, is_datetime, op, r.get("value")
            )
        if r.get("not"):
            cond = ~cond
        connector = str(r.get("connector") or "and").lower()
        if expr is None:
            expr = cond
        elif connector == "or":
            expr = expr | cond
        else:
            expr = expr & cond
    return [expr] if expr is not None else []


def _month_range(start: str, end: str) -> list[str]:
    """由 YYYY-MM-DD 起止生成 YYYY-MM 月份序列（含首尾，上限 120 个月）。"""
    try:
        y1, m1 = int(start[:4]), int(start[5:7])
        y2, m2 = int(end[:4]), int(end[5:7])
    except (ValueError, IndexError):
        return []
    months: list[str] = []
    y, m = y1, m1
    while (y, m) <= (y2, m2) and len(months) < 120:
        months.append(f"{y:04d}-{m:02d}")
        y, m = (y + 1, 1) if m == 12 else (y, m + 1)
    return months


async def _compute_plan_stats(
    session: AsyncSession, cond: list, receive_from: str, receive_to: str
) -> dict:
    """按筛选条件计算测试计划多维度统计，供 stats 端点与导出汇总共用。"""
    plan_ids_stmt = select(TestingPlan.id).where(*cond)

    by_status_rows = (
        await session.execute(
            select(TestingPlan.status, func.count(TestingPlan.id))
            .where(*cond)
            .group_by(TestingPlan.status)
        )
    ).all()
    by_status = [
        {"status": s, "name": TESTING_PLAN_STATUS.get(s, str(s)), "count": c}
        for s, c in sorted(by_status_rows)
    ]
    total_plans = sum(c for _, c in by_status_rows)
    retest_done_plans = sum(c for s, c in by_status_rows if s == PlanStatus.RETEST_DONE)
    # 初测次数：达到「初测中」及之后状态的计划各记一次初测
    first_test_count = sum(c for s, c in by_status_rows if s >= PlanStatus.TESTING)

    retest_count = (
        await session.execute(
            select(func.count(TestingPlanRetestRound.id)).where(
                TestingPlanRetestRound.plan_id.in_(plan_ids_stmt)
            )
        )
    ).scalar_one()

    # 人天统计：预估/实际总和；剩余预估人天仅统计未测试状态计划的预估人天之和
    est_mandays_total = (
        await session.execute(select(func.coalesce(func.sum(TestingPlan.est_mandays), 0.0)).where(*cond))
    ).scalar_one()
    actual_mandays_total = (
        await session.execute(select(func.coalesce(func.sum(TestingPlan.actual_mandays), 0.0)).where(*cond))
    ).scalar_one()
    remaining_est_mandays = (
        await session.execute(
            select(func.coalesce(func.sum(TestingPlan.est_mandays), 0.0)).where(
                *cond, TestingPlan.status == PlanStatus.UNTESTED
            )
        )
    ).scalar_one()

    # 按月漏洞数：筛选后计划关联漏洞按提交月份聚合（数据库无关：应用层聚合）
    months = _month_range(receive_from, receive_to) if receive_from and receive_to else []
    if not months:
        now = tznow()
        months = sorted(
            {(now.replace(day=1) - timedelta(days=30 * i)).strftime("%Y-%m") for i in range(12)}
        )
    monthly = {m: 0 for m in months}
    submit_rows = (
        await session.execute(
            select(Vul.submit_time).where(Vul.testing_plan_id.in_(plan_ids_stmt))
        )
    ).scalars().all()
    for submit_time in submit_rows:
        if submit_time is None:
            continue
        key = submit_time.strftime("%Y-%m")
        if key in monthly:
            monthly[key] += 1

    return {
        "total_plans": total_plans,
        "retest_done_plans": retest_done_plans,
        "first_test_count": first_test_count,
        "retest_count": retest_count,
        "total_test_count": first_test_count + retest_count,
        "est_mandays_total": round(float(est_mandays_total), 2),
        "actual_mandays_total": round(float(actual_mandays_total), 2),
        "remaining_est_mandays": round(float(remaining_est_mandays), 2),
        "by_status": by_status,
        "vulns_by_month": [{"month": m, "count": c} for m, c in monthly.items()],
    }


@router.get("/testing-plans", response_model=Page[TestingPlanOut])
async def list_testing_plans(
    search: str = "",
    status: int | None = None,
    test_type: str = "",
    department: str = "",
    receive_from: str = "",
    receive_to: str = "",
    filters: str = "",
    my_tests: bool = False,
    unclaimed: bool = False,
    pending: bool = False,
    sort: str = "",
    order: str = "desc",
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    user: User = Depends(require_any_perm("special:manage", "vuln:submit")),
    session: AsyncSession = Depends(get_session),
):
    """测试计划列表。my_tests 显示当前可测试系统，unclaimed 显示无人认领的测试，pending 显示待办流程。

    filters 为聚合筛选 JSON（详见 _plan_filters_condition），与上述固定参数按 AND 组合。
    """
    cond = _plan_conditions(
        search, status, test_type, department, receive_from, receive_to,
        tester_id=user.id if my_tests else None,
        unclaimed=unclaimed,
        pending=pending,
    )
    cond += _plan_filters_condition(filters)
    stmt = select(TestingPlan).where(*cond)
    stmt = apply_sort(
        stmt, TestingPlan, sort, order,
        {"id", "system_name", "plan_name", "test_type", "department", "status", "est_mandays",
         "actual_mandays", "receive_time", "first_test_done_time", "retest_done_time", "create_time"},
        TestingPlan.id.desc(),
    )
    total, items = await paginate(session, stmt, page, size)
    return Page(total=total, items=items)


@router.get("/testing-plans/stats")
async def testing_plan_stats(
    search: str = "",
    status: int | None = None,
    test_type: str = "",
    department: str = "",
    receive_from: str = "",
    receive_to: str = "",
    filters: str = "",
    pending: bool = False,
    _: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    """测试计划多维度统计：总数/复测完成数/初测次数/复测次数/总测试次数/状态分布/按月漏洞数。"""
    cond = _plan_conditions(search, status, test_type, department, receive_from, receive_to, pending=pending)
    cond += _plan_filters_condition(filters)
    return await _compute_plan_stats(session, cond, receive_from, receive_to)


PLAN_EXCEL_HEADERS = [
    "ID", "渗透测试计划名称", "测试系统", "测试类型", "所属部门",
    "工单ID", "工单提起时间", "状态", "测试人员",
    "需求接收", "初测完成", "复测通知", "复测完成",
    "预估人天", "实际人天",
    "超危数", "高危数", "中危数", "低危数", "复测轮数",
]


@router.get("/testing-plans/export")
async def export_testing_plans(
    search: str = "",
    status: int | None = None,
    test_type: str = "",
    department: str = "",
    receive_from: str = "",
    receive_to: str = "",
    filters: str = "",
    pending: bool = False,
    _: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    """导出筛选后的渗透测试计划明细与统计汇总（双 sheet Excel）。"""
    from openpyxl import Workbook

    cond = _plan_conditions(search, status, test_type, department, receive_from, receive_to, pending=pending)
    cond += _plan_filters_condition(filters)
    plans = (
        await session.execute(select(TestingPlan).where(*cond).order_by(TestingPlan.id))
    ).scalars().all()
    stats = await _compute_plan_stats(session, cond, receive_from, receive_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "渗透测试计划"
    ws.append(PLAN_EXCEL_HEADERS)
    for p in plans:
        ws.append([excel_safe(v) for v in (
            p.id, p.plan_name, p.system_name, p.test_type, p.department,
            p.ticket_id, p.ticket_time,
            TESTING_PLAN_STATUS.get(p.status, str(p.status)),
            "、".join(u.realname or u.username for u in p.testers),
            p.receive_time, p.first_test_done_time, p.retest_notice_time, p.retest_done_time,
            p.est_mandays, p.actual_mandays,
            p.stat_critical, p.stat_high, p.stat_medium, p.stat_low,
            p.retest_round_count,
        )])

    ws2 = wb.create_sheet("统计汇总")
    ws2.append(["指标", "数值"])
    ws2.append(["渗透测试计划总数", stats["total_plans"]])
    ws2.append(["复测完成计划数", stats["retest_done_plans"]])
    ws2.append(["初测次数", stats["first_test_count"]])
    ws2.append(["复测次数", stats["retest_count"]])
    ws2.append(["总测试次数（初测+复测）", stats["total_test_count"]])
    ws2.append(["预估人天总计", stats["est_mandays_total"]])
    ws2.append(["实际人天总计", stats["actual_mandays_total"]])
    ws2.append(["剩余预估人天（未测试）", stats["remaining_est_mandays"]])
    ws2.append([])
    ws2.append(["状态", "计划数"])
    for row in stats["by_status"]:
        ws2.append([row["name"], row["count"]])
    ws2.append([])
    ws2.append(["月份", "漏洞数"])
    for row in stats["vulns_by_month"]:
        ws2.append([row["month"], row["count"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "渗透测试计划导出.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


PLAN_STATUS_REVERSE = {v: k for k, v in TESTING_PLAN_STATUS.items()}


def _to_float(text) -> float:
    try:
        return float(text) if text else 0.0
    except (TypeError, ValueError):
        return 0.0


def _to_int(text: str) -> int:
    try:
        return int(float(text)) if text else 0
    except ValueError:
        return 0


@router.get("/testing-plans/import/template")
async def download_plan_import_template(_: User = Depends(require_perm("special:manage"))):
    """下载测试计划导入模板（列与导出一致，ID 留空则新增）。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "渗透测试计划"
    ws.append(PLAN_EXCEL_HEADERS)
    ws.append([
        "", "示例渗透测试计划", "示例商城系统", "渗透测试", "电商事业部",
        "", "2026-01-01", "未测试", "张三、李四",
        "2026-01-01", "", "", "",
        5, 0,
        0, 0, 0, 0, 0,
    ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "渗透测试计划导入模板.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/testing-plans/import", response_model=PlanImportResultOut)
async def import_testing_plans(
    file: UploadFile,
    user: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    """导入历史测试计划：按 ID 更新，无 ID 则新增（upsert）。测试人员按姓名/用户名匹配。"""
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "仅支持 .xlsx 格式的 Excel 文件")
    data = await file.read()
    if len(data) > 20 * 1024 * 1024:
        raise HTTPException(400, "文件大小不能超过 20MB")

    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "Excel 文件解析失败，请使用导入模板")

    # 预加载用户用于测试人员匹配（姓名优先，其次用户名）
    users = (await session.execute(select(User))).scalars().all()
    user_map: dict[str, User] = {}
    for u in users:
        if u.realname:
            user_map.setdefault(u.realname, u)
        user_map.setdefault(u.username, u)

    # 预加载工单ID占用表（手动值或自动生成值均计入），用于导入时的唯一性校验
    occupied: dict[str, int | str] = {}
    for p in (await session.execute(select(TestingPlan))).scalars().all():
        tid = p.ticket_id
        if tid:
            occupied[tid] = p.id

    ws = wb.active
    result = PlanImportResultOut()
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = [str(c).strip() if c is not None else "" for c in row]
        cells += [""] * (len(PLAN_EXCEL_HEADERS) - len(cells))
        if not any(cells):
            continue
        result.total += 1
        system_name = cells[2]
        if not system_name:
            result.failed += 1
            result.errors.append(f"第{idx}行：测试系统为必填项")
            continue

        row_id = _to_int(cells[0])
        plan = await session.get(TestingPlan, row_id) if row_id else None
        is_new = plan is None
        if is_new:
            plan = TestingPlan(creator_id=user.id)
            session.add(plan)

        plan.plan_name = cells[1]
        plan.system_name = system_name
        plan.test_type = cells[3]
        plan.department = cells[4]
        # cells[5] 工单ID：显式填写则作为手动指定值，未填写则保持原值（新记录由系统自动生成）
        plan.ticket_id_manual = cells[5] or plan.ticket_id_manual or ""
        plan.ticket_time = cells[6]
        plan.status = PLAN_STATUS_REVERSE.get(cells[7], PlanStatus.UNTESTED)
        plan.receive_time = cells[9]
        plan.first_test_done_time = cells[10]
        plan.retest_notice_time = cells[11]
        plan.retest_done_time = cells[12]
        plan.est_mandays = _to_float(cells[13])
        plan.actual_mandays = _to_float(cells[14])
        plan.stat_critical = _to_int(cells[15])
        plan.stat_high = _to_int(cells[16])
        plan.stat_medium = _to_int(cells[17])
        plan.stat_low = _to_int(cells[18])
        matched = [user_map[name] for name in cells[8].split("、") if name.strip() and name.strip() in user_map]
        if matched:
            plan.testers = matched
        # 新增或历史数据无序号时按需求接收日期自动补号
        await _assign_ticket_seq(session, plan)

        # 工单ID唯一性校验：与库中或批内其他行重复则整批终止并提示
        tid = plan.ticket_id
        if tid:
            if tid in occupied and occupied[tid] != row_id:
                raise HTTPException(400, f"第{idx}行：工单ID「{tid}」已存在，请更换后重新导入")
            occupied[tid] = "new" if is_new else row_id

        if is_new:
            result.created += 1
        else:
            result.updated += 1
    await session.commit()
    return result


# 注意：需注册在 /testing-plans/stats、/testing-plans/export、/testing-plans/import/template 等静态路径之后，防止路径吞噬
@router.get("/testing-plans/{row_id}", response_model=TestingPlanOut)
async def get_testing_plan(
    row_id: int,
    _: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    """单条计划详情（含测试人员/关联漏洞/关联报告/复测轮次），供流程抽屉刷新。"""
    return await get_or_404(session, TestingPlan, row_id, "渗透测试计划不存在")


async def _assign_ticket_seq(session: AsyncSession, row: TestingPlan) -> None:
    """按需求接收日期分配当日「最大编号+1」的录入次序（ticket_seq）。

    委托给 ticket_service.assign_ticket_seq：与测试计划同源的工单ID分配逻辑，
    现为测试计划 / 非渗透计划两表共享同一当日序号序列（见 services/ticket_service.py）。
    """
    await ticket_service.assign_ticket_seq(session, row)


async def _check_ticket_id_unique(
    session: AsyncSession, ticket_id: str, exclude_id: int | None = None,
    linked_nonpen_ids: list[int] | None = None,
) -> None:
    """校验工单ID唯一性：与「显示编号」口径一致——手动指定值本身，或纯自动记录
    （ticket_id_manual 为空）由 receive_time+ticket_seq 生成的值，均不得与其他记录重复。

    委托给 ticket_service.check_ticket_id_unique：现为测试计划 / 非渗透计划两表全局唯一。
    排除自身及联动创建的非渗透记录（联动双方共享同一工单ID，须相互排除）。
    """
    excludes = [(TestingPlan, exclude_id)] if exclude_id is not None else []
    for nid in (linked_nonpen_ids or []):
        excludes.append((NonpenPlan, nid))
    await ticket_service.check_ticket_id_unique(
        session, ticket_id, exclude=excludes or None,
    )


@router.post("/testing-plans", response_model=TestingPlanOut)
async def create_testing_plan(
    body: TestingPlanIn,
    user: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    data = body.model_dump()
    create_nonpen = bool(data.pop("create_nonpen", False))
    nonpen_test_items = list(data.pop("nonpen_test_items") or [])
    row = TestingPlan(**data, creator_id=user.id)
    await _assign_ticket_seq(session, row)
    await _check_ticket_id_unique(session, row.ticket_id)
    session.add(row)
    # 勾选「创建非渗透」：同一工单联动生成非渗透计划（共享工单ID与接收日期）
    if create_nonpen:
        if not nonpen_test_items:
            raise HTTPException(400, "已勾选「创建非渗透」，请至少选择一个非渗透测试项")
        for k in nonpen_test_items:
            if k not in NONPEN_ITEMS:
                raise HTTPException(400, f"不支持的测试项：{k}")
        if not row.receive_time and not row.ticket_id_manual:
            raise HTTPException(400, "已勾选「创建非渗透」，请填写「需求接收日期」（用于生成共享工单ID）或手动指定工单ID")
        await session.flush()  # 先持久化测试计划拿到 id，供非渗透记录引用
        session.add(NonpenPlan(
            plan_name=row.plan_name,
            system_name=row.system_name,
            test_type=row.test_type,
            department=row.department,
            ticket_time=row.ticket_time,
            receive_time=row.receive_time,
            ticket_seq=row.ticket_seq,  # 混合工单：非渗透复用测试计划的当日序号
            ticket_id_manual=row.ticket_id_manual,
            asset_ids=list(row.asset_ids or []),
            items=nonpen_service.build_items(nonpen_test_items),
            testing_plan_id=row.id,  # 标记联动来源，用于双向同步与级联删除
            detail=row.detail,
            creator_id=user.id,
        ))
    await session.commit()
    await session.refresh(row)
    return row


@router.post("/testing-plans/{row_id}/claim", response_model=TestingPlanOut)
async def claim_testing_plan(
    row_id: int,
    user: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    """认领测试计划：当前用户加入测试人员（幂等）；未测试状态自动进入初测中。"""
    row = await get_or_404(session, TestingPlan, row_id, "渗透测试计划不存在")
    if all(u.id != user.id for u in row.testers):
        row.testers.append(user)
    if row.status == 10:
        row.status = 20
    await session.commit()
    await session.refresh(row)
    return row


@router.post("/testing-plans/{row_id}/quit", response_model=TestingPlanOut)
async def quit_testing_plan(
    row_id: int,
    user: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    """退出认领：当前用户移出测试人员列表。"""
    row = await get_or_404(session, TestingPlan, row_id, "渗透测试计划不存在")
    row.testers = [u for u in row.testers if u.id != user.id]
    await session.commit()
    await session.refresh(row)
    return row


class CompleteNoVulnIn(BaseModel):
    conclusion: str = ""  # 测试结论：记录到计划并写入无漏洞报告，可留空使用默认结论
    generate_report: bool = True  # 是否同步生成「未发现安全漏洞」报告草稿
    title: str = ""  # 报告标题：留空时自动生成「yyyymmdd+测试系统+渗透测试报告（无漏洞）」


def _no_vul_section_html(system_name: str, conclusion: str) -> str:
    """无漏洞报告的「测试结论」章节 HTML（供报告编辑器继续编辑与 Word 导出）。"""
    parts = [
        "<p><strong>测试状态：</strong>初测</p>",
        "<p><strong>测试结论：</strong>经本次安全测试，"
        f"未发现{html_mod.escape(system_name or '被测系统')}存在安全漏洞，安全测试通过。</p>",
    ]
    if conclusion.strip():
        escaped = "<br/>".join(
            html_mod.escape(line) for line in conclusion.strip().splitlines()
        )
        parts.append(f"<p><strong>补充说明：</strong></p><p>{escaped}</p>")
    return "".join(parts)


def _no_vul_mandays(test_start: str, test_end: str) -> float:
    """无漏洞报告实际人天：与报告模块口径一致，结束日期 - 开始日期 + 1，非法时 0。"""
    try:
        start = datetime.strptime(test_start, "%Y-%m-%d")
        end = datetime.strptime(test_end, "%Y-%m-%d")
    except (TypeError, ValueError):
        return 0.0
    if end < start:
        return 0.0
    return float((end - start).days + 1)


async def _create_no_vul_report(
    session: AsyncSession, plan: TestingPlan, user: User, title: str, conclusion: str,
) -> Report:
    """为无漏洞计划生成「安全测试通过」报告草稿：单章节测试结论，无漏洞关联。

    标题同计划范围内查重，重复时自动追加「-1」「-2」后缀；测试周期默认取
    需求接收日期至确认当天（未填接收日期时取当天），实际人天自动计算。
    标题不含「复测」，按初测报告口径计入计划实际人天。
    """
    today = tznow().date().isoformat()
    candidate = title
    suffix = 0
    while (
        await session.execute(
            select(Report.id).where(
                Report.title == candidate, Report.testing_plan_id == plan.id
            ).limit(1)
        )
    ).scalar_one_or_none() is not None:
        suffix += 1
        candidate = f"{title}-{suffix}"
    test_start = plan.receive_time or today
    report = Report(
        title=candidate,
        project_name=plan.system_name,
        author=user.realname or user.username,
        test_start=test_start,
        test_end=today,
        status="draft",
        testing_plan_id=plan.id,
        creator_id=user.id,
    )
    report.sections.append(ReportSection(
        order=0, title="测试结论",
        content_html=_no_vul_section_html(plan.system_name, conclusion),
    ))
    session.add(report)
    await session.flush()
    report.actual_mandays = _no_vul_mandays(report.test_start, report.test_end)
    return report


@router.post("/testing-plans/{row_id}/complete-no-vuln", response_model=TestingPlanOut)
async def complete_plan_no_vuln(
    row_id: int,
    body: CompleteNoVulnIn,
    user: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    """无漏洞闭环：测试完成且确认未发现安全漏洞时，计划直接流转为「测试通过」。

    - 前置校验：计划无关联漏洞（有漏洞须走整改/复测链路）、状态允许流转且未重复确认；
    - 数据记录：状态置为测试通过、初测完成时间打点、记录无漏洞测试结论；
    - 报告归档：默认同步生成「未发现安全漏洞」报告草稿（可导出 Word/PDF 归档）；
    - 通知：站内信告知测试人员与计划创建人；
    - 后续若补录/关联新漏洞，计划自动重开为「初测中」（见 reopen_passed_plan）。
    """
    plan = await get_or_404(session, TestingPlan, row_id, "渗透测试计划不存在")
    if not plan_service.can_operate(user, plan):
        raise HTTPException(403, "仅认领者或管理员可确认无漏洞完结")
    if plan.status == PlanStatus.PASSED:
        raise HTTPException(400, "该计划已确认无漏洞（测试通过），无需重复操作")
    if not vuln_service.can_plan_transition(plan.status, PlanStatus.PASSED):
        raise HTTPException(400, "当前状态不允许确认无漏洞完结")
    vul_count = (
        await session.execute(
            select(func.count(Vul.id)).where(Vul.testing_plan_id == row_id)
        )
    ).scalar_one()
    if vul_count:
        raise HTTPException(400, "该计划存在关联漏洞，不能确认无漏洞，请先处理漏洞后走复测流程")

    # 状态流转与数据记录
    plan.status = PlanStatus.PASSED
    plan.no_vul_conclusion = body.conclusion.strip()
    if not plan.first_test_done_time:
        plan.first_test_done_time = tznow().date().isoformat()

    # 报告归档：生成无漏洞报告草稿（标题不含「复测」，计入计划实际人天）
    report: Report | None = None
    if body.generate_report:
        title = body.title.strip() or (
            f"{tznow().date().strftime('%Y%m%d')}{plan.system_name}渗透测试报告（无漏洞）"
        )
        report = await _create_no_vul_report(session, plan, user, title, body.conclusion)
        # 计划的 reports 关系在创建报告前已预加载，需重新加载后再重算实际人天，避免遍历时漏掉新报告
        await session.refresh(plan, attribute_names=["reports"])
        await plan_service.refresh_mandays(session, plan.id)

    # 站内信通知：测试人员与计划创建人（去重，排除操作人本人）
    notice_ids = {u.id for u in plan.testers}
    if plan.creator_id:
        notice_ids.add(plan.creator_id)
    notice_ids.discard(user.id)
    report_hint = f"，已生成报告《{report.title}》" if report is not None else ""
    for uid in notice_ids:
        session.add(Message(
            user_id=uid,
            msg_type="plan",
            title=f"测试计划「{plan.system_name}」已确认无漏洞",
            content=(
                f"{user.realname or user.username} 确认该计划测试完成且未发现安全漏洞，"
                f"状态流转为「测试通过」{report_hint}"
            ),
        ))
    await session.commit()
    await session.refresh(plan)
    return plan


@router.post("/testing-plans/{row_id}/attach-vulns", response_model=TestingPlanOut)
async def attach_vulns_to_plan(
    row_id: int,
    body: dict,
    user: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    """从漏洞库批量关联漏洞到当前测试计划（需求1）。

    属录入漏洞阶段：仅已认领该计划的账号可操作（管理员未认领不放行）；
    已关联其他计划的漏洞会被转移至当前计划。
    """
    row = await get_or_404(session, TestingPlan, row_id, "渗透测试计划不存在")
    if not plan_service.is_plan_claimant(user, row):
        raise HTTPException(403, "仅已认领该测试计划的账号可关联漏洞")
    vul_ids = [int(i) for i in (body.get("vul_ids") or [])]
    if not vul_ids:
        raise HTTPException(400, "请选择要关联的漏洞")
    vulns = await _load_vulns(session, vul_ids)
    for v in vulns:
        v.testing_plan_id = row_id
    # 已确认无漏洞（测试通过）的计划重新关联到漏洞时自动重开为「初测中」
    await plan_service.reopen_passed_plan(session, row_id)
    await plan_service.refresh_stats(session, row_id)
    await session.commit()
    await session.refresh(row)
    return row


@router.put("/testing-plans/{row_id}", response_model=TestingPlanOut)
async def update_testing_plan(
    row_id: int,
    body: TestingPlanIn,
    user: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    row = await get_or_404(session, TestingPlan, row_id, "渗透测试计划不存在")
    if body.status != row.status and not plan_service.can_operate(user, row):
        raise HTTPException(403, "仅认领者或管理员可修改测试状态")
    # 校验状态流转合法性（仅当状态有变化时）
    if body.status != row.status and not vuln_service.can_plan_transition(row.status, body.status):
        raise HTTPException(400, f"不允许从当前状态流转到目标状态")
    # 编辑页直接流转为「测试通过」时，同样要求计划无关联漏洞（与无漏洞完结接口口径一致）
    if body.status == PlanStatus.PASSED and row.status != PlanStatus.PASSED:
        vul_count = (
            await session.execute(
                select(func.count(Vul.id)).where(Vul.testing_plan_id == row_id)
            )
        ).scalar_one()
        if vul_count:
            raise HTTPException(400, "该计划存在关联漏洞，不能流转为「测试通过」")
    # 手动流转到「复测中」时记一轮复测（已有进行中轮次则不重复计数）
    if body.status == 50 and row.status != 50:
        plan_service.start_retest_round(session, row, "手动流转至复测中", user.id)
    data = body.model_dump()
    # 联动相关字段仅创建时生效：编辑不回写 create_nonpen 勾选，也不存在 nonpen_test_items 列
    data.pop("create_nonpen", None)
    data.pop("nonpen_test_items", None)
    for k, v in data.items():
        setattr(row, k, v)
    # 补生成工单ID序号（历史/导入数据无序号时自动补齐）
    await _assign_ticket_seq(session, row)
    # 联动非渗透计划：编辑测试计划公共字段时双向同步；联动双方共享同一工单ID，唯一性校验需相互排除
    linked = (await session.execute(
        select(NonpenPlan).where(NonpenPlan.testing_plan_id == row.id)
    )).scalars().all()
    # 保存前校验工单ID唯一性（手动指定值或自动生成值均不可重复）
    await _check_ticket_id_unique(
        session, row.ticket_id, exclude_id=row.id,
        linked_nonpen_ids=[np.id for np in linked],
    )
    # 有关联漏洞时统计以自动重算为准，覆盖手填值
    if row.vuls:
        await plan_service.refresh_stats(session, row.id)
    # 有关联初测报告时实际人天自动计算（仅纳入初测报告，复测报告不计入）
    await plan_service.refresh_mandays(session, row.id)
    # 联动双向同步：编辑测试计划公共字段时，自动同步更新其联动的非渗透计划
    for np in linked:
        nonpen_service.sync_linked_fields(row, np)
    await session.commit()
    await session.refresh(row)
    return row


@router.delete("/testing-plans/{row_id}")
async def delete_testing_plan(
    row_id: int,
    _: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    row = await session.get(TestingPlan, row_id)
    if row:
        # 级联删除联动创建的非渗透计划（联动双向：删除任一方，另一方同步删除）
        linked = (await session.execute(
            select(NonpenPlan).where(NonpenPlan.testing_plan_id == row_id)
        )).scalars().all()
        for np in linked:
            await session.delete(np)
        # 解除漏洞与报告的关联，避免悬挂外键
        await session.execute(
            update(Vul).where(Vul.testing_plan_id == row_id).values(testing_plan_id=None)
        )
        await session.execute(
            update(Report).where(Report.testing_plan_id == row_id).values(testing_plan_id=None)
        )
        await session.delete(row)
        await session.commit()
    return {"msg": "删除成功"}


# ---------- 春耕行动 ----------
async def _load_vulns(session: AsyncSession, vul_ids: list[int]) -> list[Vul]:
    if not vul_ids:
        return []
    vulns = (await session.execute(select(Vul).where(Vul.id.in_(vul_ids)))).scalars().all()
    if len(vulns) != len(set(vul_ids)):
        raise HTTPException(400, "部分漏洞不存在")
    return list(vulns)


@router.get("/spring-actions", response_model=Page[SpringActionOut])
async def list_spring_actions(
    search: str = "",
    sort: str = "",
    order: str = "desc",
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    _: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    cond = []
    if search:
        cond.append(
            SpringAction.report_no.ilike(f"%{search}%")
            | SpringAction.system_name.ilike(f"%{search}%")
            | SpringAction.doc_no.ilike(f"%{search}%")
        )
    stmt = select(SpringAction).where(*cond)
    stmt = apply_sort(
        stmt, SpringAction, sort, order,
        {"id", "report_no", "system_name", "year", "phase", "appeal_success",
         "score_deduction", "doc_no", "create_time"},
        SpringAction.id.desc(),
    )
    total, items = await paginate(session, stmt, page, size)
    return Page(total=total, items=items)


@router.post("/spring-actions", response_model=SpringActionOut)
async def create_spring_action(
    body: SpringActionIn,
    user: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    row = SpringAction(**body.model_dump(exclude={"vul_ids"}), creator_id=user.id)
    row.vuls = await _load_vulns(session, body.vul_ids)
    session.add(row)
    await session.commit()
    await session.refresh(row)
    return row


@router.put("/spring-actions/{row_id}", response_model=SpringActionOut)
async def update_spring_action(
    row_id: int,
    body: SpringActionIn,
    _: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    row = await get_or_404(session, SpringAction, row_id, "春耕行动记录不存在")
    for k, v in body.model_dump(exclude={"vul_ids"}).items():
        setattr(row, k, v)
    row.vuls = await _load_vulns(session, body.vul_ids)
    await session.commit()
    await session.refresh(row)
    return row


@router.delete("/spring-actions/{row_id}")
async def delete_spring_action(
    row_id: int,
    _: User = Depends(require_perm("special:manage")),
    session: AsyncSession = Depends(get_session),
):
    row = await session.get(SpringAction, row_id)
    if row:
        await session.delete(row)
        await session.commit()
    return {"msg": "删除成功"}
