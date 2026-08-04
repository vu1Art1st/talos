from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import cast, func, select, String
from sqlalchemy.ext.asyncio import AsyncSession

from app.constants import ASSET_STATUS, URL_TAG
from app.core.deps import get_current_user, require_perm
from app.core.query import get_or_404, paginate, apply_sort
from app.core.sanitize import excel_safe
from app.db import get_session
from app.models import Asset, Group, GroupMember, User, vuln_assets
from app.schemas import AssetImportResultOut, AssetIn, AssetOut, Page

router = APIRouter(prefix="/assets", tags=["资产"])

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Excel 列定义（顺序即模板列顺序）
EXCEL_HEADERS = [
    "系统命名*", "子系统名称", "部门", "系统类型", "公网URL", "内网URL",
    "开放端口与服务", "中间件", "数据库", "系统负责人", "状态", "备注",
]

STATUS_REVERSE = {v: k for k, v in ASSET_STATUS.items()}
URL_TAG_REVERSE = {v: k for k, v in URL_TAG.items()}


def _asset_conditions(search: str, department: str, status: int | None) -> list:
    cond = []
    if search:
        like = f"%{search}%"
        cond.append(
            Asset.name.ilike(like)
            | Asset.sub_system.ilike(like)
            | Asset.department.ilike(like)
            | cast(Asset.public_urls, String).ilike(like)
            | cast(Asset.internal_urls, String).ilike(like)
        )
    if department:
        cond.append(Asset.department.ilike(f"%{department}%"))
    if status is not None:
        cond.append(Asset.status == status)
    return cond


@router.get("", response_model=Page[AssetOut])
async def list_assets(
    search: str = "",
    department: str = "",
    status: int | None = None,
    sort: str = "",
    order: str = "desc",
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=200),
    _: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    cond = _asset_conditions(search, department, status)
    stmt = select(Asset).where(*cond)
    stmt = apply_sort(
        stmt, Asset, sort, order,
        {"id", "name", "sub_system", "department", "system_type", "status", "create_time"},
        Asset.id.desc(),
    )
    total, items = await paginate(session, stmt, page, size)
    return Page(total=total, items=items)


@router.post("", response_model=AssetOut)
async def create_asset(
    body: AssetIn,
    _: User = Depends(require_perm("asset:manage")),
    session: AsyncSession = Depends(get_session),
):
    asset = Asset(**body.model_dump())
    session.add(asset)
    await _sync_owners_to_group_members(
        session, body.department, [o.model_dump() for o in body.owners],
    )
    await session.commit()
    await session.refresh(asset)
    return asset


# ---------- Excel 导入导出（需在 /{asset_id} 之前注册，避免路径冲突） ----------
def _dump_owners(owners: list | None) -> str:
    return ";".join(
        "/".join(filter(None, [o.get("name", ""), o.get("phone", ""), o.get("email", "")]))
        for o in (owners or [])
    )


def _dump_public_urls(urls: list | None) -> str:
    return ";".join(
        f"{u.get('url', '')}|{URL_TAG.get(u.get('tag', 10), '互联网')}" for u in (urls or [])
    )


def _parse_owners(text: str) -> list[dict]:
    owners = []
    for part in filter(None, (p.strip() for p in text.split(";"))):
        fields = part.split("/")
        owners.append({
            "name": fields[0].strip(),
            "phone": fields[1].strip() if len(fields) > 1 else "",
            "email": fields[2].strip() if len(fields) > 2 else "",
        })
    return owners


async def _sync_owners_to_group_members(
    session: AsyncSession, department: str, owners: list[dict],
) -> None:
    """录入的系统负责人自动同步到组织管理（GroupMember）。

    以资产部门作为归属组织；同名负责人已存在时跳过，保证幂等。
    """
    if not department or not owners:
        return
    group = (
        await session.execute(select(Group).where(Group.name == department))
    ).scalar_one_or_none()
    if group is None:
        return
    existing = set((await session.execute(select(GroupMember.name))).scalars().all())
    for o in owners:
        name = (o.get("name") or "").strip()
        if name and name not in existing:
            session.add(GroupMember(
                group_id=group.id,
                name=name,
                phone=(o.get("phone") or "").strip(),
                email=(o.get("email") or "").strip(),
            ))
            existing.add(name)


def _parse_public_urls(text: str) -> list[dict]:
    urls = []
    for part in filter(None, (p.strip() for p in text.split(";"))):
        url, _, tag_name = part.partition("|")
        urls.append({"url": url.strip(), "tag": URL_TAG_REVERSE.get(tag_name.strip(), 10)})
    return urls


def _parse_list(text: str) -> list[str]:
    return [p.strip() for p in text.split(";") if p.strip()]


def _dump_port_services(items: list | None) -> str:
    return ";".join(
        f"{i.get('port', '')}:{i.get('service', '')}".strip(":") for i in (items or [])
    )


def _parse_port_services(text: str) -> list[dict]:
    """解析「[端口]:[服务]」对，分号分隔多组，如 80:Web服务;443:HTTPS。"""
    items = []
    for part in filter(None, (p.strip() for p in text.split(";"))):
        port, _, service = part.partition(":")
        items.append({"port": port.strip(), "service": service.strip()})
    return items


def _dump_name_versions(items: list | None) -> str:
    return ";".join(
        "/".join(filter(None, [i.get("name", ""), i.get("version", "")])) for i in (items or [])
    )


def _parse_name_versions(text: str) -> list[dict]:
    """解析「名称/版本」条目，分号分隔多组，如 Nginx/1.24;Tomcat/9.0。"""
    items = []
    for part in filter(None, (p.strip() for p in text.split(";"))):
        name, _, version = part.partition("/")
        items.append({"name": name.strip(), "version": version.strip()})
    return items


def _build_workbook(assets: list[Asset]):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "资产"
    ws.append(EXCEL_HEADERS)
    for a in assets:
        ws.append([excel_safe(v) for v in (
            a.name, a.sub_system, a.department, a.system_type,
            _dump_public_urls(a.public_urls),
            ";".join(a.internal_urls or []),
            _dump_port_services(a.port_services),
            _dump_name_versions(a.middlewares),
            _dump_name_versions(a.databases),
            _dump_owners(a.owners),
            ASSET_STATUS.get(a.status, "线上"),
            a.remark,
        )])
    return wb


def _xlsx_response(wb, filename: str) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.get("/export")
async def export_assets(
    search: str = "",
    department: str = "",
    status: int | None = None,
    _: User = Depends(require_perm("asset:manage")),
    session: AsyncSession = Depends(get_session),
):
    cond = _asset_conditions(search, department, status)
    assets = (
        await session.execute(select(Asset).where(*cond).order_by(Asset.id))
    ).scalars().all()
    return _xlsx_response(_build_workbook(list(assets)), "资产导出.xlsx")


@router.get("/import/template")
async def download_import_template(_: User = Depends(require_perm("asset:manage"))):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "资产"
    ws.append(EXCEL_HEADERS)
    ws.append([
        "示例商城系统", "订单中心", "电商事业部", "自有系统（正式）",
        "https://shop.example.com|互联网;https://oa-shop.example.com|办公网",
        "http://10.0.0.8:8080", "80:Web服务;443:HTTPS;8080:管理后台",
        "Nginx/1.24;Tomcat/9.0", "MySQL/8.0",
        "张三/13800000000/zhangsan@example.com;李四/13900000000/lisi@example.com",
        "线上", "示例行，导入前请删除",
    ])
    return _xlsx_response(wb, "资产导入模板.xlsx")


@router.post("/import", response_model=AssetImportResultOut)
async def import_assets(
    file: UploadFile,
    _: User = Depends(require_perm("asset:manage")),
    session: AsyncSession = Depends(get_session),
):
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "仅支持 .xlsx 格式的 Excel 文件")
    data = await file.read()
    if len(data) > 20 * 1024 * 1024:
        raise HTTPException(400, "文件大小不能超过 20MB")

    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "Excel 文件解析失败，请使用导入模板")

    ws = wb.active
    result = AssetImportResultOut()
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = [str(c).strip() if c is not None else "" for c in row]
        cells += [""] * (len(EXCEL_HEADERS) - len(cells))
        if not any(cells):
            continue
        result.total += 1
        name = cells[0]
        if not name:
            result.failed += 1
            result.errors.append(f"第{idx}行：系统命名为必填项")
            continue
        owners = _parse_owners(cells[9])
        session.add(Asset(
            name=name,
            sub_system=cells[1],
            department=cells[2],
            system_type=cells[3],
            public_urls=_parse_public_urls(cells[4]),
            internal_urls=_parse_list(cells[5]),
            port_services=_parse_port_services(cells[6]),
            middlewares=_parse_name_versions(cells[7]),
            databases=_parse_name_versions(cells[8]),
            owners=owners,
            status=STATUS_REVERSE.get(cells[10], 10),
            remark=cells[11],
        ))
        await _sync_owners_to_group_members(session, cells[2], owners)
        result.success += 1
    await session.commit()
    return result


@router.get("/{asset_id}", response_model=AssetOut)
async def get_asset(
    asset_id: int,
    _: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    asset = await get_or_404(session, Asset, asset_id, "资产不存在")
    return asset


@router.put("/{asset_id}", response_model=AssetOut)
async def update_asset(
    asset_id: int,
    body: AssetIn,
    _: User = Depends(require_perm("asset:manage")),
    session: AsyncSession = Depends(get_session),
):
    asset = await get_or_404(session, Asset, asset_id, "资产不存在")
    for k, v in body.model_dump().items():
        setattr(asset, k, v)
    await _sync_owners_to_group_members(
        session, body.department, [o.model_dump() for o in body.owners],
    )
    await session.commit()
    await session.refresh(asset)
    return asset


@router.delete("/{asset_id}")
async def delete_asset(
    asset_id: int,
    _: User = Depends(require_perm("asset:manage")),
    session: AsyncSession = Depends(get_session),
):
    used = (
        await session.execute(
            select(func.count()).select_from(vuln_assets).where(vuln_assets.c.asset_id == asset_id)
        )
    ).scalar_one()
    if used:
        raise HTTPException(400, "该资产已关联漏洞记录，无法删除")
    asset = await session.get(Asset, asset_id)
    if asset:
        await session.delete(asset)
        await session.commit()
    return {"msg": "删除成功"}
